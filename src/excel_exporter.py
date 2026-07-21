"""Export weld extraction results to Excel (.xlsx)."""

from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

from .weld_extractor import WeldFeature, WeldExtractionResult


HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
DATA_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def export_to_excel(result: WeldExtractionResult, output_path: str):
    """Export extraction results to Excel workbook.
    
    Sheets:
      1. Weld List — each row is one weld with all features
      2. Part Info — base material thicknesses
      3. Summary — model metadata
    """
    wb = Workbook()
    
    # ── Sheet 1: Weld List ──
    ws1 = wb.active
    ws1.title = "Weld List"
    
    headers = [
        'Weld ID', 'Name', 'Source',
        'Center X (mm)', 'Center Y (mm)', 'Center Z (mm)',
        'Length (mm)', 'Leg Length (mm)', 'Throat (mm)',
        'Part 1 ID', 'Part 2 ID',
        'T1 (mm)', 'T2 (mm)',
        'Joint Type', 'Gap (mm)',
        'Process Type',
        'Distance to Prev Weld (mm)',
        'Notes'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    
    # Write data
    for row_idx, weld in enumerate(result.welds, 2):
        data = [
            weld.id, weld.name, weld.source,
            round(weld.center[0], 3), round(weld.center[1], 3), round(weld.center[2], 3),
            round(weld.length, 2), round(weld.leg_length, 2), round(weld.throat_thickness, 2),
            weld.part1_id, weld.part2_id,
            round(weld.T1, 2), round(weld.T2, 2),
            weld.joint_type, round(weld.gap, 3),
            weld.process_type,
            round(weld.distance_to_prev_weld, 2),
            ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws1.cell(row=row_idx, column=col, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center')
    
    # Auto-width
    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = max(12, len(headers[col-1]) + 2)
    
    # Freeze header row
    ws1.freeze_panes = 'A2'
    
    # ── Sheet 2: Part Info ──
    ws2 = wb.create_sheet("Part Info")
    
    ws2.cell(row=1, column=1, value='Part ID').font = HEADER_FONT
    ws2.cell(row=1, column=1).fill = HEADER_FILL
    ws2.cell(row=1, column=1).border = THIN_BORDER
    ws2.cell(row=1, column=2, value='Thickness (mm)').font = HEADER_FONT
    ws2.cell(row=1, column=2).fill = HEADER_FILL
    ws2.cell(row=1, column=2).border = THIN_BORDER
    
    row_idx = 2
    for pid, thickness in sorted(result.part_thicknesses.items()):
        ws2.cell(row=row_idx, column=1, value=pid).font = DATA_FONT
        ws2.cell(row=row_idx, column=1).border = THIN_BORDER
        ws2.cell(row=row_idx, column=2, value=thickness).font = DATA_FONT
        ws2.cell(row=row_idx, column=2).border = THIN_BORDER
        row_idx += 1
    
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15
    
    # ── Sheet 3: Summary ──
    ws3 = wb.create_sheet("Summary")
    
    summary_data = [
        ('Model Name', result.model_name),
        ('Total Parts', result.total_parts),
        ('Total Welds', result.total_welds),
        ('', ''),
        ('Notes', ''),
    ]
    for note in result.notes:
        summary_data.append(('', note))
    
    for row_idx, (key, value) in enumerate(summary_data, 1):
        ws3.cell(row=row_idx, column=1, value=key).font = Font(bold=True, size=10)
        ws3.cell(row=row_idx, column=1).border = THIN_BORDER
        ws3.cell(row=row_idx, column=2, value=value).font = DATA_FONT
        ws3.cell(row=row_idx, column=2).border = THIN_BORDER
    
    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 60
    
    # Save
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    
    return str(output)
